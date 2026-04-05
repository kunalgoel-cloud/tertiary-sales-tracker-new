"""
deals_promos_module.py
──────────────────────
Deals & Promos tab for the Mamanourish Executive Tracker.

Supported channels:
  • BigBasket  — generates a date-range CSV with state columns
  • Amazon     — fills SVD PRICE and BAU PRICE into the embedded deal sheet template

Both channels show a live Channel Performance panel above the price entry form
so the user can see revenue, units, and avg price per SKU before deciding on prices.
"""

import streamlit as st
import pandas as pd
import calendar
import io
import base64
from datetime import date, timedelta

import plotly.express as px
from openpyxl import load_workbook
from openpyxl.styles import Font

# ──────────────────────────────────────────────────────────────
# AMAZON DEAL SHEET TEMPLATE  (embedded — no upload needed)
# ──────────────────────────────────────────────────────────────
_AMZN_TEMPLATE_B64 = (
    "UEsDBBQACAgIANUyfVwAAAAAAAAAAAAAAAAYAAAAeGwvZHJhd2luZ3MvZHJhd2luZzEueG1sndBdb"
    "sIwDAfwE+wOVd5pWhgTQxRe0E4wDuAlbhuRj8oOo9x+0Uo2aXsBHm3LP/nvzW50tvhEYhN8I+qy"
    "EgV6FbTxXSMO72+zlSg4gtdgg8dGXJDFbvu0GTWtz7ynIu17XqeyEX2Mw1pKVj064DIM6NO0DeQg"
    "ppI6qQnOSXZWzqvqRfJACJp7xLifJuLqwQOaA+Pz/k3XhLY1CvdBnRz6OCGEFmL6Bfdm4KypB65R"
    "PVD8AcZ/gjOKAoc2liq46ynZSEL9PAk4/hr13chSvsrVX8jdFMcBHU/DLLlDesiHsSZevpNlRnfu"
    "gbdoAx2By8i4OPjj3bEqyTa1KCtssV7ercyzIrdfUEsHCAdiaYMFAQAABwMAAFBLAwQUAAgICADV"
    "Mn1cAAAAAAAAAAAAAAAAGAAAAHhsL3dvcmtzaGVldHMvc2hlZXQxLnhtbKXd23IbxxHG8SfIO7Bw"
    "5SRlAjPTM9OtIpmKLMuSD5Itxck1TIISyiTAAqCD3z4gTlzMLHf/snUhicPeRWP7AwT8ROye/evz"
    "7c3Jx8liOZ3PzgfudDQ4mcwu51fT2bvzwa//ef61Dk6Wq/Hsanwzn03OB39MloN/Xfzt7NN88fvy"
    "/WSyOlnvYLY8H7xfre6eDIfLy/eT2/HydH43ma2/cz1f3I5X6y8X74bLu8VkfLXZ6PZm6EejNLwd"
    "T2eD7R6eLMg+5tfX08vJs/nlh9vJbLXdyWJyM16t21++n94t93u7/Vzt7nZ6uZgv59er08v57W5P"
    "6w4uh5PPl5NNQ3rU0O0l6eh2vPj9w93X613erbv4bXozXf2x6euwm4/ngw+L2ZPdPr4+tHG/zZP1"
    "7T/5eHuzL/7shPVdHUwb2lH3n138c3tyo6Fzxa5kXB8L3tb48rCnW7abw0R2Ebk42+zy58XF2d34"
    "3eTtZPXr3c+L4cXZ8LC++ct/p5NPy8bfT+5j+tt8/vv9Fy+vzgejwWGjZu3zzUB/Xpxcfliu5rcv"
    "JtN371frh8Pg5GpyPf5ws/pmfvO/6dXq/XrNn6ZwWH8z/3QojqebvV/Ob5ab33c72283OLmdzrZ/"
    "jj9v/vx02KPqbsv2bfxuG/+wzeg0x85twm6bcNjGx1PfvY3stpGH23Gnobu3uNsmHraJcjgQ7Zuk"
    "3SbpYZNTFzo3ybtN8mETuR9D1ya620QPm9hp99273RbWPMzdd8WN9uMcNTrr2eYQAddorecAuH0G"
    "3EMIct8N7TPgQvMe9YTN7VPgjmLQ198+B/5hqrob0XD7kNg80p6NV+OLs8X808lis/H9Iyf5+2BW"
    "j731rd0X/XtdtdzUrleX69WPF6Oz4cf73e4qntYV7rjim7rCH1c8qyvCccW3dYUcVzyvK+JxxXd1"
    "RTqueLGt8I2KfFzxsq7Q44rv61ux44ofWo5YcVB/bCkpjupP25LQLCkO66uWkofjOlzn4BAGvw2D"
    "k84w+M3+pLm/YgpPtyWxWVKM4ZuWkmIOz1pKikF821JSTOL5tiRt87R+Ai4O8nfb7+dtIsXK77/Y"
    "38TF2fXmVt6PF5Orwclicn3/zScv1k9Vy+nm+Hzn//GV+/pH//ez4fW2WzktEvxyuzdt3dtL/+Tl"
    "bm9+cPHCD79y//z+YW+pOD7fb/dl28fj6ag4wj88NL6+Y6dulGxk9m1xBH883knR7k9HB8dXB+dV"
    "z8F7vT/4w93Cz+XCL+XCm3Lh7f5+DIu0BpTWUKe1eCQ+DX1pbSkp09pSUqa1LvHFAXseGmmVOKrS"
    "GpoH3Oq0hva0bhM63N6kVakM7ancJnG4PWjrrba/ihSG7hQGlMLQmcKjO611CnsOyutQprBc+KVc"
    "eFMuvG0sHKVQUAqlSqEvnsmfSn8KW0rKFLaUlCmsS3zxT8ZzaaTQ5zqF2++vX3Dto1HGUHpj6EId"
    "Q+mPYVi/9t/8Ku7V99IdQ0ExlM4YNo9Ky51+1fP911LGsFz4pVx4Uy683d+PMoYRxTDWMSzu5NPY"
    "H8OWkjKGLSVlDOsSX76Yiz0xjH0xjH8qhvEvxDB2xzCiGMbOGMbGs53LdQxjTwxjGcNy4Zdy4U25"
    "8DY+EsOEYpi2czt6L1AcqaepP4ctJWUOW0rKHO6aab6Y98Vunu9qwvYNhdRJ3BXI9lVmy6vI1JtE"
    "L/WrxdSbRP9oElN3EhNKYupMYt+9ftVRcBSbjGKTW2JTNPw098empaSMTUtJGZvcEpvyrceuJj4a"
    "m9wXm/ynYpP/Qmxyd2wyik3ujE3fvX7VUXAUG0Wx0ZbYlG8BtD82LSVlbFpKythoHZtQvgfY1Wyf"
    "wVPLmwBtHh9piY32xia4OjbaH5tH3wVod2wUxUY7Y9N3r191FBzFxlBsbLu33JxU+Zq9pabKTVtN"
    "GZy2mjI5bf2Ur9t3Nbo9Ar5Ozq7Ado8CqZJj4P1jPC2e6F4aeP9Y6oV1J8ZQYqwzMX339lVfwWt7"
    "5F3fvTV3ZGj4UOhQ2NxOFEeHl1hP66Vv6qVn9dK3R0vHXTPgc1tx8a7RzG7JN5rZLYVGM7slaTTT"
    "XDpuhvmNC3UzoW4m1M2Euplq6af9UjwsvTpaOm6Zvdl3UrcsdctStyx1y/Lo8WNv+Vysm4l1M7Fu"
    "JtbNxEebYS/8XaqbSXUzqW4m1c2kR5thLyddrpvJdTO5bibXzeRHm2EvUpzWzWjdjNbNaN2MPtoM"
    "+6fPWd2M1c1Y3YzVzdhjzfju59B9M/dly81/ah2a2S/lh2b2S41m9kuNZo6Wjpthz9N+99yqjWZc"
    "3Yyrm3F1M+7RZuB/xPi6GV83Uz9P+/p52j/6PO27n6cf/qnz3c+OjcLuZ65GYfezSqOw+xHfKOx+"
    "tDYKux8pD4WBvhII3QlrFHZPv1FIJxPoZAKdTKCTCXQygU4m0MkInYzQyQidjNDJCJ2M0MkInYzQ"
    "yQidjNDJRDqZSCcT6WQinUykk4l0MpFOJtLJRDqZSCeT6GQSnUyik0l0MolOJtHJJDqZRCeT6GQS"
    "nUymk8l0MplOJtPJZDqZTCeT6WQynUymk8l0Mkono3QySiejdDJKJ6N0Mkono3QySiejdDJGJ2N0"
    "MkYnY3QyRidjdDJGJ2N0MkYnY3QyboQBbURn40Z0OG5Ep+NGdDxuROfjRnRAbkQn5EZ0RG6EZ/QF"
    "yIln1EOQzUo8ox6Wa1biGfWAVrMSz6iHgpqVeEY9iNKsxDPq4YdmJZ4RxgGHdcBhHnDYBxwGAoel"
    "wGEicNgIHEYCh5XAYSZw2AkchgKHpcBhKnDYChzGAoe1wGEucNgLHAYDh8XAYTJw2AwcRgOH1cBh"
    "NnDYDRyGA4flwGE6cNgOHMYDh/XAYT5w2A8cBgSHBcFhQnDYEBxGBIcVwWFGcNgRHIYEhyXBYUpw"
    "2BIcxgSHNcFhTnDYExwGBYdFwWFScNgUHEYFh1XBYVZw2BUchgWHZcFhWnDYFhzGBYd1wWFecNgX"
    "HAYGh4XBYWJw2BgcRgaHlcFhZnDYGTx2Bo+dwWNn8NgZPHYGj53BY2fw2Bk8dgaPncFjZ/DYGTx2"
    "Bo+dwWNn8NgZPHYGj53BY2fw2Bk8dgaPncFjZ/D8ZxD4DyFgZ/D8xxD4zyHwH0TAzuCxM3jsDB47"
    "g8fO4LEzeOwMHjuDx87gsTN47AweO4PHzuCxM3jsDB47g8fO4LEzeOwMHjuDx87gsTN47AweO4PH"
    "zuCxM3jsDB47g8fO4LEzeOwMHjuDx87gsTN47AweO4PHzuCxM3jsDB47g8fO4LEzeOwMHjuDx87g"
    "sTN47AweO4PHzuCxM3jsDB47g8fO4LEzeOwMHjuDx87gsTN47AwBO0PAzhCwMwTsDAE7Q8DOELADC"
    "DYGwQ7g2BnEOwMgp1BsDMIdgbBziDYGQQ7g2BnEOwMgp1BsDMIdgbBziDYGQQ7g2BnEOwMgp1BsD"
    "MIdgbBziDYGQQ7g2BnEOwMgp1BsDMIdgbBziDYGQQ7g2BnEOwMgp1BsDMIdgbBziDYGQQ7g2BnEO"
    "wMgp1BsDMIdgbBziDYGQQ7g2BnEOwMgp1BsDMIdgbBziDYGQQ7g2BnEOwMgp1BsDMIdgbBziDYGQ"
    "Q7g2BnEOwMgp1BsDMIdgbBziDYGQQ7g2BnEOwMgp1BsDMIdgbBziDYGQQ7g2BnEOwMgp1BsDMIdg"
    "bBziDYGQQ7g2BnEOwMgp1BsDMIdgbBziDYGQQ7g2BnEOwMgp1BsDMIdgbBziDYGQQ7g2BnEOwMgp"
    "1BsDMIdgbBziDYGQQ7g2BnEOwMgp1BsDMIdgbBziDYGQQ7g2BnEOwMgp1BsDMIdgbBziDYGQQ7g2"
    "BnEOwMgp1BsDMIdgbBziDYGQQ7g2BnEOwMgp1BsDMIdgbBziDYGQQ7g2BnEOwMgp1BsDMIdgbBzi"
    "DYGQQsTNE7AwRO0PEzhCxM0TsDBE7Q8TOELEzROwMETtDxM4QsTNE7AwRO0PEzhCxM0TsDBE7Q8TO"
    "ELEzROwMETtDxM4QsTNE7AwRO0PEzhCxM0TsDBE7Q8TOELEzROwMETtDxM4QsTNE7AwRO0PEzhCx"
    "M0TsDBE7Q8TOELEzROwMETtDxM4QsTNE7AwRO0PEzhCxM0TsDBE7Q8TOELEzROwMETtDxM4QsTNE"
    "flZHflpHfl5HfmJHfmbHLzi1I54RP7kjP7sjP70jdoaInSFiZ4jYGSJ2hoidIWJniNgZInaGiJ0h"
    "YmeI2BkidoaInSFiZ4jYGSJ2hoidIWJniNgZInaGiJ0hYmeI2BkidoaInSFiZ4jYGSJ2hoidIWJn"
    "iNgZInaGiJ0hYmeI2BkidoaInSFiZ4jYGSJ2hoSdIWFnSNgZEnaGhJ0hYWdI2BkSdoaEnSFhZ0jY"
    "GRJ2hoSdIWFnSNgZEnaGhJ0hYWdI2BkSdoaEnSFhZ0jYGRJ2hoSdIWFnSNgZEnaGhJ0hYWdI2BkS"
    "doaEnSFhZ0jYGRJ2hoSdIWFnSNgZEnaGhJ0hYWdI2BkSdoaEnSFhZ0jYGRJ2hoSdIWFnSNgZEnae"
    "hJ0hYWdI2BkSdoaEnSFhZ0jYGRJ2hoSdIWFnSNgZEnaGhJ0hYWdI2BkSdoaEnSFhZ0jYGRK/jgS/"
    "kAS/kgS/lAS/lgS/mMQXXE0Cz4hfT4JfUAI7Q8LOkLAzJOwMCTtDws6QsDMk7AwJO0PCzpCwMyTs"
    "DAk7Q8LOkLAzJOwMCTtDws6QsDMk7AwJO0PCzpCwMyTsDAk7Q8LOkLAzJOwMCTtDws6QsDNk7AwZ"
    "O0PGzpCxM2TsDBk7Q8bOkLEzZOwMGTtDxs6QsTNk7AwZO0PGzpCxM2TsDBk7Q8bOkLEzZOwMGTtD"
    "xs6QsTNk7AwZO0PGzpCxM2TsDBk7Q8bOkLEzZOwMGTtDxs6QsTNk7AwZO0PGzpCxM2TsDBk7Q8bO"
    "kLEzZOwMGTtDxs6QsTNk7AwZO0PGzpCxM2TsDBk7Q8bOkLEzZOwMGTtDxs6QsTNk7AwZO0PGzpCx"
    "M2TsDBk7Q8bOkLEzZOwMGTtDxs6QsTNk7AwZO0PGzpCxM2TsDBk7Q8bOkLEzZOwMmV+5kl+6kl+7"
    "kl+8kl++kl+/8gsuYIlnxC9hiZ0hY2fI2BkydoaMnSFjZ8jYGTJ2hoydIWNnyNgZMnaGjJ0hY2fI"
    "2BkydoaMnSFjZ8jYGTJ2hoydQbEzKHYGxc6g2BkUO4NiZ1DsDIqdQbEzKHYGxc6g2BkUO4NiZ1Ds"
    "DIqdQbEzKHYGxc6g2BkUO4NiZ1DsDIqdQbEzKHYGxc6g2BkUO4NiZ1DsDIqdQbEzKHYGxc6g2BkU"
    "O4NiZ1DsDIqdQbEzKHYGxc6g2BkUO4NiZ1DsDIqdQbEzKHYGxc6g2BkUO4NiZ1DsDIqdQbEzKHYG"
    "xc6g2BkUO4NiZ1DsDIqdQbEzKHYGxc6g2BkUO4NiZ1DsDIqdQbEzKHYGxc6g2BkUO4NiZ1DsDIqd"
    "QbEzKHYGxc6g2BkUO4NiZ1DsDIqdQbEzKHYGxc6g2BkUO4NiZ1DsDIqdQbEzKHYGxc6g2BkUO4Ni"
    "Z1DsDIqdQbEzKHYGxc6g2BkMO4NhZzDsDIadwbAzGHYGw85g2BkMO4NhZzDsDIadwbAzGHYGw85g"
    "2BkMO4NhZzDsDIadwbAzGHYGw85g2BkMO4NhZzDsDIadwbAzGHYGw85g2BkMO4NhZzDsDIadwbAz"
    "GHYGw85g2BkMO4NhZzDsDIadwbAzGHYGw85g2BkMO4NhZzDsDIadwbAzGHYGw85g2BkMO4NhZzDs"
    "DIadwbAzGHYGw85g2BkMO4NhZzDsDIadwbAzGHYGw85g2BkMO4NhZzDsDIadwbAzGHYGw85g2BkM"
    "O4NhZzDsDIadwbAzGHYGw85g2BkMO4NhZzDsDIadwbAzGHYGw85g2BkMO4NhZzDsDIadwbAzGHYG"
    "w85g2BkMO4P1O8Nw+X4yWT0br8YXZ3eL6Wz1+m41nc+W62/djd9Nfhov3k1ny5Pf5qv1lueD0en9"
    "w/16Pl9NFvdf3U9hMr46fHEzuV5tqgYni+2tbP6+mt/ttt3t9+1k9eHuZL6YTmar8f0Nng9uxrOr"
    "5eX4bnJfc7UYf5rO3p0snkyvzgeLl1fbZj/NF79vGr74P1BLBwghXZ8E2REAAPrCAABQSwMEFAAI"
    "CAGA1TJ9XAAAAAAAAAAAAAAAACMAAAB4bC93b3Jrc2hlZXRzL19yZWxzL3NoZWV0MS54bWwucmVsc43"
    "PSwrCMBAG4BN4hzB7k9aFiDTtRoRupR5gSKYPbB4k8dHbm42i4MLlzM98w181DzOzG4U4OSuh5AUw"
    "ssrpyQ4Szt1xvQMWE1qNs7MkYaEITb2qTjRjyjdxnHxkGbFRwpiS3wsR1UgGI3eebE56FwymPIZB"
    "eFQXHEhsimIrwqcB9ZfJWi0htLoE1i2e/rFd30+KDk5dDdn044XQAe+5WCYxDJQkcP7avcOSZxZE"
    "XYmvivUTUEsHCK2o602zAAAAKgEAAFBLAwQUAAgICADVMn1cAAAAAAAAAAAAAAAAEQAAAGRvY1Byb3"
    "BzL2NvcmUueG1sbZHfToMwFIefwHcgvYcW5lQaYBeaJSaamAwz411TjtCM/klbZXt7C9vQ6O7a/L"
    "7z9ZzTYrWXffQF1gmtSpQmBEWguG6Eakv0Wq/jOxQ5z1TDeq2gRAdwaFVdFdxQri28WG3AegEuCi"
    "LlKDcl6rw3FGPHO5DMJYFQIfzQVjIfrrbFhvEdawFnhNxgCZ41zDM8CmMzG9FJ2fBZaT5tPwkajq"
    "EHCco7nCYp/mE9WOkuFkzJL1IKfzBwET2HM713YgaHYUiGxYSG/lP89vy0mUaNhRpXxQFVxakRyi"
    "0wD00UBPT43DnZLu4f6jWqMpItY5LHaV6TW5rmdHn9XuA/9aPweNa2ety5Tvho07Gw0RGdkwL/+5"
    "fqG1BLBwinGAOsFQEAAOMBAABQSwMEFAAICAgA1TJ9XAAAAAAAAAAAAAAAABMAAAB4bC90aGVtZS90"
    "aGVtZTEueG1szVfJbtswEP2C/gOhe6PN2ozYQdLY6KELULfomZGopaFIgaSb5u9LUbJFbU7QOEV0"
    "MTl8M/M4M+TQl1d/Sgx+I8YLSlaGfWEZAJGYJgXJVsaP79v3oQG4gCSBmBK0Mh4RN67W7y7hUuSo"
    "RECqE76EKyMXolqaJo+lGPILWiEi11LKSijklGVmwuCDNFti07Es3yxhQYxWnz1Hn6ZpEaNbGu9L"
    "RERBR+8jSBiH4zQ/MbA50iQhlOh5wQpEQSa4Vl4qplyH4KpsiAL14j0hanSyQsEoSAUV0tPGhZHG"
    "ty0cMezG3OZaKO8eSSmebU/BpOxjhAtpG6HDdd8JXOVBjpMQxHGv+Lx3c4coSSRZAlsL1C8+HwFq"
    "nvMqjC42O1Nio+GtJIhuxXnRovhRxpJQ+po258A1RIHJL40lj03FK8vDeW5guVlTDqYYt2E0JeOJ"
    "/iXRvL2JEpSMLI/xXX06hQBJQxusIXX1USodxz8vBoO6o6WORm/Fj9EyEXJk9qVKzT6ommBViAKK"
    "M+0VW1EI9MRsJSujDe3QKJBWwpYWxdgo0CJvaWBGWmpaC9bSwg0uCm5CZcFMP8WUPliQf9mrBr6B"
    "mmWeG/N7ZGPmWbWqoknCssinLGZVBeU5nd+a7OEMOxjXFAtXs0zq7pzzmt/+oJvjWdaSQRSgqtOz"
    "O7s5ff5YV4WxmkjCH4ciJrqom9NKk9QuQycf8J4lyod4pqsknmULdP1Guv62dH8IR2od/YmQ5MW0"
    "W3425+p8UzMAyyOkuFriknJNYn8j//4u+RvMQwreX0t4lbqwFfbZp2HvIFGK3XAbJQ/EaGm/6Eqx"
    "3zYgrnY/ZQmWcXGf7BimwcE97iMM1Q6xbRzOOi+6xi2hnRrnX38JbnUebt6sj2DF7m9G68O9uxkt"
    "Ef5s1RcVPaaMODqxvkJA2yyjdqSbr+4Xu7i6NxwLh6awgz8Le2R3VBHqcJxspdjlnlJqf4y/r2W8"
    "6jawx9l/JF92Lv/in6tPotQt+rB80qg9rLx7VnltfSNln/tCcG/v2xR4yZRQTbjre/diYjBHs7fH"
    "Etf79j4e/QdQSwcIG4aB/DADAADXDgAAUEsDBBQACAgIANUyfVwAAAAAAAAAAAAAAAAUAAAAeGwvc2"
    "hhcmVkU3RyaW5ncy54bWzdllFz2hoQhT+B34HsPVLmTgiRMHEyChiIBwiRMHHydpYW6VrpjtydxN"
    "Dpj9+eABtGkEkmfejkSaPdu7v//7ZvpfabL1kK96g0l6LjnL1oOIAilBEXcce5mQ9qlw5ow0TEUi"
    "mw46xRO2+6v7S1NkBLhe44iTGrV/W6DhPMmH4hVyjIs5QqY4ZeVVzXK4Us0gmiyUL6eaNxUc8YFw"
    "6EMhem47ReOpAL/jnHfmloNp1uW/Nu23R7wdppu266bbt9L20LFJlU0JcRHnFNWVZxuYpk7Bvn3K"
    "SVyIk/2zf1goqJZRpJAYF44CYBw758PULm5lDQvGoaBRXJE6ZiLvatw8KDmf+2f1U5uXdzxNHoLx"
    "bn8w9+c9/zu3857+0br4RBtVJcUz/AhDGYK37PDMKYZNxgBeeEZUzInBYk+65xz/Nu3J4PnlovVc"
    "4NXAlW8BpcpDf/AVIKPS65wAiCPmYKfWXZ6bc1zhRoVHcvv0UaOkKUmWduNYFDsNM2N4oZa2G5GEY"
    "8eDckm+BAsPEoX/hoED9SKaI8qSnfN4hhpSXmiR/rssTNaAHIJZ3DSKpO8hWYLhopl cMXC5LeDo"
    "F+6F4M/jmqvwVCK6AnBmEVRbnO6W4Nl9y3Z0ek2LjktkzmtCDKeZzvCTqn9H2BId8qmannOMYmLB"
    "RLjeQLDGFVOGlvkYyvsnhWMWGraDa8Kj2pTHiYFRLogESRexRTCwVwkfYCJNgmoXykUJI850lYJ3"
    "/X78+f20dZTChAuunyBYjeBCS/2/a4mtMsF3u4GsS/5tDXDyWP/WgUp7n/puazC7Paqx L0Mp6MY"
    "XGdR2NRa6ZizN4B1PT6tqHkU+1177pqbdgtiAvCtBll4af/am6A2bbYY7eO4KPCdnLViR+EOyR6P"
    "L5mLe/x+U1u4KAf8bS8TFvj9Y6WbjYKVvF+6nuf8zVnor/Pyx5nStzw9BuPA/Ls5aw59mqFvFje+"
    "Y6t782jv0sdsO7A+JTFHLzHZjmqJ51oiFmL8rRAzTnHIkRzG3fRbzbYIu+1Oxr867ADX91dADo+f"
    "Tv6CwgWSlkEK2Yt/RsH7KvE4/bt1/AVQSwcIWxpzaSYDAAD2CQAAUEsDBBQACAgIANUyfVwAAAAA"
    "AAAAAAAAAAA0AAAB4bC9zdHlsZXMueG1s1VjLcpswFP2C/gOje+SRR4GMbaezdCetutJMuvXKIIO"
    "mHh5JJJB/r0TZiQMGjAOEsAC97rlH517doBvczBj1nrBURPAQ+Kc94GGeilHhI9D8HcYn1wBT2nER"
    "ohqjgOARzrMBN9C1Qek7xwwRj7RkErkIw0Tq/hlClE8yQOhU55qYnE5IhbapyDFUuMRopa8Qo7Pd"
    "6l5AhwoFDuJ755yhdwmEklUKJTJ+mgkGRZSTFy0gDOIAorZDYMkwDHYbk4zQ/MbA50iQhlOh5wQpE"
    "QSa4Vl4qplyH4KpsiAL14j0hanSyQsEoSAUV0tPGhZHGty0cMezG3OZaKO8eSSmebU/BpOxjhAtp"
    "G6HDdd8JXOVBjpMQxHGv+Lx3c4coSSRZAlsL1C8+HwFqnvMqjC42O1Nio+GtJIhuxXnRovhRxpJQ"
    "+po258A1RIHJL40lj03FK8vDeW5guVlTDqYYt2E0JeOJ/iXRvL2JEpSMLI/xXX06hQBJQxusIXX1"
    "USodxz8vBoO6o6WORm/Fj9EyEXJk9qVKzT6ommBViAKKM+0VW1EI9MRsJSujDe3QKJBWwpYWxdgo"
    "0CJvaWBGWmpaC9bSwg0uCm5CZcFMP8WUPliQf9mrBr6BmmWeG/N7ZGPmWbWqoknCssinLGZVBeU5"
    "nd+a7OEMOxjXFAtXs0zq7pzzmt/+oJvjWdaSQRSgqtOzO7s5ff5YV4WxmkjCH4ciJrqom9NKk9Qu"
    "Qycf8J4lyod4pqsknmULdP1Guv62dH8IR2od/YmQ5MW0W3425+p8UzMAyyOkuFriknJNYn8j//4u"
    "+RvMQwreX0t4lbqwFfbZp2HvIFGK3XAbJQ/EaGm/6Eqx3zYgrnY/ZQmWcXGf7BimwcE97iMM1Q6x"
    "bRzOOi+6xi2hnRrnX38JbnUebt6sj2DF7m9G68O9uxktEf5s1RcVPaaMODqxvkJA2yyjdqSbr+4X"
    "u7i6NxwLh6awgz8Le2R3VBHqcJxspdjlnlJqf4y/r2W86jawx9l/JF92Lv/in6tPotQt+rB80qg9"
    "rLx7VnltfSNln/tCcG/v2xR4yZRQTbjre/diYjBHs7fHEtf79j4e/QdQSwcIEyVm/y4DAABkFwAA"
    "UEsDBBQACAgIANUyfVwAAAAAAAAAAAAAAAAPAAAAeGwvd29ya2Jvb2sueG1snZPRbpswFIafYO+AfJ"
    "8YorRrUUi1JduUtlqrNOnWq8kxhlhgH2abhO7pd+IAWpubaDcYc/Dnzz+HyU2jymAnjJWgExINQ"
    "xIIzSGVOk/IevV1cEUC65hOWQlaJORVWHIz/TDZgyk2AEWA67VNyNa5KqbU8q1QzA6hEhorGRjF"
    "HE5NTm1lBEvtVginSjoKw0uqmNTkSIjNOQzIMsnFHHithHZHiBElc2hvt7KyHU01JzgluQELmRty"
    "UK0IDTgVDRde6OqNkOLnGClmiroaILJCi40spXv1Xj1ml5Da6LhlDHqNw5oY9493quxebqLxed4n"
    "YV7T6zf2TXTxf6QopFH0DjVmp1mcr8V4T1LnYfov0rbItG+3R0OnE8+37XjoToeNuZNWbkpBAs0U"
    "TpZ3wY+H5f381/zLbHDL9OgC+/jw/iLFNieBiSXemEU6JkikHTIVmdQi/Y4Mi885K7nfUjTu3jo/"
    "BrWRCfkGkJfiyS+b1daBmjPHno8/0giNc4jtu2qbQg59CtzXALDQBZJ7sE/jH0u0N1Dr1BlZHVCz"
    "reCFrTHPj7/Zeknp02KvP/MV/WlmeXX7cLlnIU2j5ll+0uviz93jy8q+JP6oeITj1R+IdslO/wJQ"
    "SwcIdab/5tMBAAALBAAAUEsDBBQACAgIANUyfVwAAAAAAAAAAAAAAAAaAAAAeGwvX3JlbHMvd29ya2"
    "Jvb2sueG1sLnJlbHOtkk1OwzAQhU/AHSzvGyflRwjV6QYhdQvlAMaeOFFiT2RPgdweQ0Waoine0Z"
    "X1njXvffJ4s/10HXuHEBv0khdZzhl4jabxVvLX/dPqnrNIyhvVoQfJB4h8W15tnqFTlGZi3fSRpR"
    "AfJa+J+gchoq7BqZhhDz7dVBicoiSDFb3SrbIg1nl+J8I0g5dnmWxnJA87U3C2H3r4TzZWVaPhEf"
    "XBgaeZCkFpFlKgChZI8h95NIsshXExz7C+JEOkoUtvOEIc9VL99UXraxXAvFBIC55STO0lmJtLwn"
    "xgaGMNQCeQ0fpGTcfiYm7/wOhDJHS/SBbRdpBpdDO1b4itA1JGkTq1j04qFGdfvPwCUEsHCP9Evb"
    "gJAQAAKgMAAFBLAwQUAAgICADVMn1cAAAAAAAAAAAAAAAACwAAAF9yZWxzLy5yZWxzpZBNasMwEEZP"
    "0DuI2cfjZFFKiZxNKWQXinuAqTS2hS2NkJQ2uX1FobSGLApdzs/3eDP7w8Uv6p1TdhI0bJsWFAcj"
    "1oVRw2v/vHkAlQsFS4sE1nDlDIfubv/CC5WayZOLWVVIyBqmUuIjYjYTe8qNRA51MkjyVGqZRoxk"
    "ZhoZd217j+k3A7oVUx2thnS0W1D9NfL/2Oi5kKVCaCTxJqaaTsXVU1RPaeSiwYo51Xb+2mgqGfC2"
    "0O7vQjIMzvCTmLPnUG55rTd+bC4Lfkia30Tmbxdcfbz7BFBLBwhYIhhl1gAAALkBAABQSwMEFAAI"
    "CAGA1TJ9XAAAAAAAAAAAAAAAAAsAAAB4bC9tZXRhZGF0YeNiNtQzkBLhYjQU4jK0NDMwtDAxNjRWeMq"
    "uIeVVwcWamhfv7iQk7JibWpSZnKjvk18c75iXnpqTWqzA5CDikRIUwsXCwSDBKMTi5+/nKsXm5B8S"
    "4u+rxOEf5hrk5uMfrsXjWFCSX6zgl1hUlF9uwG3B4MDgwRDAEMGQxMHBIMAswaDAnMXOwSTw//9/"
    "9ioWDmYJxhmMDABQSwcIccSVQJcAAACVAAAAUEsDBBQACAgIANUyfVwAAAAAAAAAAAAAAAATAAAAW0Nvbn"
    "RlbnRfVHlwZXNdLnhtbLVU207DMAz9Av6hyitas/GAEFrHA7BHQAI+wGvcNVqbRLF36d/jtgOJMdi4"
    "vTRJj32Oj+tmfLWpq2SFkax3mRqlQ5Wgy72xbp6p56fp4EIlxOAMVN5hphokdTU5GT81ASmRZEeZ"
    "KpnDpdaUl1gDpT6gE6TwsQaWY5zrAPkC5qjPhsNznXvH6HjALYeajG+wgGXFyXX/vqXOFIRQ2RxY"
    "6tJCppLbjYB9me1ZH5G3cmanmIEvCpuj8fmylpTUz4olSTSaqZC8E/GGufipzNZvGrHqYqi0gU53"
    "fQhKrcK9fIBoDf7GCYWIYKhE5LpK1z4uun2v+QCR76AWUr2p9BtIultG6bahh+uYWQex2SWskcEA"
    "w/94oRIimkeOMpO0z8+7gOO9HK7DRFgL5z7NLUSvm2/08ItxyX3EQYiCRrb40axU9iAo6TbwL53ud"
    "Jybao962+oO+UtlljsD90l1QP/8XXMPDJesaQ3WffanzLxfvOrr7tqbvABQSwcILqmfZ3UBAAA2"
    "BQAAUEsBAhQAFAAICAgA1TJ9XAdiaYMFAQAABwMAABgAAAAAAAAAAAAAAAAAAAAAAHhsL2RyYXdpbmdzL"
    "2RyYXdpbmcxLnhtbFBLAQIUABQACAgIANUyfVwhXZ8E2REAAPrCAAAYAAAAAAAAAAAAAAAAAEsBAAB4"
    "bC93b3Jrc2hlZXRzL3NoZWV0MS54bWxQSwECFAAUAAgICADVMn1crajrTbMAAAAqAQAAIwAAAAAAAA"
    "AAAAAAAAAqEwAAeGwvd29ya3NoZWV0cy9fcmVscy9zaGVldDEueG1sLnJlbHNQSwECFAAUAAgICADV"
    "Mn1cpxgDrBUBAADjAQAAEQAAAAAAAAAAAAAAAABuFAAAZG9jUHJvcHMvY29yZS54bWxQSwECFAAUAAgI"
    "CADVMn1cG4aB/DADAADXDgAAEwAAAAAAAAAAAAAAAADCFQAAeGwvdGhlbWUvdGhlbWUxLnhtbFBLAQIU"
    "ABQACAgIANUyfVxbGnNpJgMAAPYJAAAUAAAAAAAAAAAAAAAAADMZAAB4bC9zaGFyZWRTdHJpbmdzLnht"
    "bFBLAQIUABQACAgIANUyfVwTJWb/LgMAAGQXAAANAAAAAAAAAAAAAAAAAJscAAB4bC9zdHlsZXMueG1s"
    "UEsBAhQAFAAICAgA1TJ9XHWm/+bTAQAACwQAAA8AAAAAAAAAAAAAAAAABCAAAHhsL3dvcmtib29rLnht"
    "bFBLAQIUABQACAgIANUyfVz/RL24CQEAACoDAAAaAAAAAAAAAAAAAAAAABQiAAB4bC9fcmVscy93b3Jr"
    "Ym9vay54bWwucmVsc1BLAQIUABQACAgIANUyfVxYIhhl1gAAALkBAAALAAAAAAAAAAAAAAAAAGUjAABf"
    "cmVscy8ucmVsc1BLAQIUABQACAgIANUyfVxxxJVAlwAAAJUAAAALAAAAAAAAAAAAAAAAAHQkAAB4bC9t"
    "ZXRhZGF0YVBLAQIUABQACAgIANUyfVwuqZ9ndQEAADYFAAATAAAAAAAAAAAAAAAAAEQlAABbQ29udGVu"
    "dF9UeXBlc10ueG1sUEsFBgAAAAAMAAwAEgMAAPomAAAAAA=="
)

AMZN_TEMPLATE_BYTES = base64.b64decode(_AMZN_TEMPLATE_B64)

# ──────────────────────────────────────────────────────────────
# BIG BASKET FORMAT CONSTANTS
# ──────────────────────────────────────────────────────────────
BB_HEADERS = [
    "Code", "Product Description",
    "Start Date (DD-MM-YYYY)", "End Date (DD-MM-YYYY)",
    "Discount Type", "Discount Value",
    "Redemption Limit - Qty Per Campaign", "Pan India",
    "ANDHRA PRADESH", "TELANGANA", "ASSAM", "BIHAR", "CHHATTISGARH",
    "GUJARAT", "HARYANA_DELHI&GURGAON", "JHARKHAND", "KARNATAKA",
    "KERALA", "MADHYA PRADESH", "MAHARASHTRA - Mumbai", "MAHARASHTRA - Pune",
    "ORISSA", "PUNJAB", "RAJASTHAN", "TAMIL NADU",
    "UTTAR PRADESH_Noida", "WEST BENGAL",
]
BB_STATE_COLUMNS = BB_HEADERS[8:]

DEFAULT_CITY_MAP = {
    "Mumbai-DC":     "MAHARASHTRA - Mumbai",
    "Pune-DC":       "MAHARASHTRA - Pune",
    "Bangalore-DC":  "KARNATAKA",
    "Bangalore-DC2": "KARNATAKA",
    "Hyderabad-DC":  "TELANGANA",
    "Kolkata-DC":    "WEST BENGAL",
    "Chennai-DC":    "TAMIL NADU",
    "Ahmedabad-DC":  "GUJARAT",
    "Delhi-DC":      "HARYANA_DELHI&GURGAON",
    "Gurgaon-DC":    "HARYANA_DELHI&GURGAON",
}

TIER_LABELS = ["BAU", "SVD", "Weekend", "Liq"]
TIER_HELP = {
    "BAU":     "Business-as-usual: standard everyday price.",
    "SVD":     "Slow-mover deal: used on days 1–10 of month for overstocked / slow SKUs.",
    "Weekend": "Weekend push: used Sat–Sun for SKUs that under-index on weekdays.",
    "Liq":     "Liquidation: deepest price for very slow-moving SKUs.",
}

AMZN_COL_SVD        = 13
AMZN_COL_BAU        = 14
AMZN_DATA_START_ROW = 2


# ──────────────────────────────────────────────────────────────
# SHARED SIGNAL HELPERS
# ──────────────────────────────────────────────────────────────

def _compute_signals_with_city(chan_hist: pd.DataFrame, lookback_days: int = 30) -> pd.DataFrame:
    if chan_hist.empty:
        return pd.DataFrame(columns=["channel_sku", "location", "str", "doc"])
    h = chan_hist.copy()
    h["date_dt"] = pd.to_datetime(h["date"], errors="coerce")
    cutoff = h["date_dt"].max() - timedelta(days=lookback_days - 1)
    h = h[h["date_dt"] >= cutoff]
    if h.empty:
        return pd.DataFrame(columns=["channel_sku", "location", "str", "doc"])
    rows = []
    for (sku, city), g in h.groupby(["item_name", "city"]):
        days_w = g[g["qty_sold"] > 0]["date_dt"].nunique()
        rows.append({"channel_sku": sku, "location": city,
                     "str": round(days_w / lookback_days, 4),
                     "_avg": g["qty_sold"].sum() / lookback_days})
    sig = pd.DataFrame(rows)
    mv = sig["_avg"].max()
    sig["doc"] = ((1 - sig["_avg"] / mv) * 100).round(1) if mv > 0 else 50.0
    return sig.drop(columns=["_avg"])


def _compute_signals_national(chan_hist: pd.DataFrame, lookback_days: int = 30) -> pd.DataFrame:
    if chan_hist.empty:
        return pd.DataFrame(columns=["channel_sku", "str", "doc"])
    h = chan_hist.copy()
    h["date_dt"] = pd.to_datetime(h["date"], errors="coerce")
    cutoff = h["date_dt"].max() - timedelta(days=lookback_days - 1)
    h = h[h["date_dt"] >= cutoff]
    if h.empty:
        return pd.DataFrame(columns=["channel_sku", "str", "doc"])
    rows = []
    for sku, g in h.groupby("item_name"):
        days_w = g[g["qty_sold"] > 0]["date_dt"].nunique()
        rows.append({"channel_sku": sku,
                     "str": round(days_w / lookback_days, 4),
                     "_avg": g["qty_sold"].sum() / lookback_days})
    sig = pd.DataFrame(rows)
    mv = sig["_avg"].max()
    sig["doc"] = ((1 - sig["_avg"] / mv) * 100).round(1) if mv > 0 else 50.0
    return sig.drop(columns=["_avg"])


def _classify(str_val, doc_val, is_svd_day, is_weekend, doc_threshold=70.0):
    if str_val < 0.20 and doc_val > doc_threshold:
        return "Liq"
    if str_val >= 0.20 and doc_val > doc_threshold:
        return "SVD" if is_svd_day else "BAU"
    if str_val < 0.20 and doc_val <= doc_threshold:
        return "Weekend" if is_weekend else "BAU"
    return "BAU"


# ──────────────────────────────────────────────────────────────
# CHANNEL PERFORMANCE PANEL  (shared widget, used by both channels)
# ──────────────────────────────────────────────────────────────

def _render_channel_performance(chan_hist: pd.DataFrame, channel_label: str, lookback_days: int):
    """
    Renders a compact performance panel for the given channel's history:
      - Summary metrics (total revenue, units, avg price)
      - SKU-level bar chart ranked by revenue
      - SKU × metric table for reference while setting prices
    """
    with st.expander(f"📈 {channel_label} Channel Performance — last {lookback_days} days",
                     expanded=True):
        if chan_hist.empty:
            st.info(f"No {channel_label} sales data available.")
            return

        h = chan_hist.copy()
        h["date_dt"] = pd.to_datetime(h["date"], errors="coerce")
        cutoff = h["date_dt"].max() - timedelta(days=lookback_days - 1)
        h = h[h["date_dt"] >= cutoff]

        if h.empty:
            st.info(f"No data in the last {lookback_days} days.")
            return

        # Summary metrics
        total_rev  = h["revenue"].sum()
        total_qty  = h["qty_sold"].sum()
        avg_price  = total_rev / total_qty if total_qty > 0 else 0
        days_range = max((h["date_dt"].max() - h["date_dt"].min()).days + 1, 1)
        drr        = total_rev / days_range

        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Revenue (₹)", f"₹{total_rev:,.0f}")
        m2.metric("Units Sold",  f"{total_qty:,.0f}")
        m3.metric("Avg Price",   f"₹{avg_price:,.1f}")
        m4.metric("Daily Run Rate", f"₹{drr:,.0f}")

        # SKU-level breakdown
        sku_df = (
            h.groupby("item_name")
             .agg(revenue=("revenue", "sum"), qty=("qty_sold", "sum"))
             .reset_index()
             .sort_values("revenue", ascending=False)
        )
        sku_df["avg_price"] = (sku_df["revenue"] / sku_df["qty"].where(sku_df["qty"] > 0)).round(1)
        sku_df["rev_share"] = (sku_df["revenue"] / sku_df["revenue"].sum() * 100).round(1)

        fig = px.bar(
            sku_df.sort_values("revenue"),
            x="revenue", y="item_name", orientation="h",
            color="rev_share", color_continuous_scale="Blues",
            text=sku_df.sort_values("revenue")["revenue"].apply(lambda v: f"₹{v:,.0f}"),
            height=max(250, len(sku_df) * 42),
            labels={"revenue": "Revenue (₹)", "item_name": ""},
        )
        fig.update_traces(textposition="outside")
        fig.update_layout(coloraxis_showscale=False, margin=dict(l=10, r=80, t=10, b=10))
        st.plotly_chart(fig, use_container_width=True)

        sku_df.columns = ["SKU", "Revenue (₹)", "Units", "Avg Price (₹)", "Rev Share %"]
        st.dataframe(
            sku_df.style.format({
                "Revenue (₹)": "₹{:,.0f}", "Units": "{:,.0f}",
                "Avg Price (₹)": "₹{:.1f}", "Rev Share %": "{:.1f}%",
            }),
            hide_index=True, use_container_width=True,
        )


# ──────────────────────────────────────────────────────────────
# SIGNALS PANEL
# ──────────────────────────────────────────────────────────────

def _signals_expander(sig_df: pd.DataFrame, doc_thresh: float, has_city: bool = True):
    with st.expander("🔬 View STR & DOC Signals", expanded=False):
        st.caption(
            "**STR** = fraction of days with ≥1 sale in the lookback window. "
            "**DOC Index** = 0 (fastest) → 100 (slowest). "
            "Together they drive the suggested tier for each SKU."
        )
        d = sig_df.copy()
        d["Suggested Tier"] = d.apply(
            lambda r: _classify(r["str"], r["doc"], True, False, doc_thresh), axis=1)
        rn = {"channel_sku": "SKU", "str": "STR", "doc": "DOC Index"}
        if has_city:
            rn["location"] = "City"
        d.rename(columns=rn, inplace=True)
        st.dataframe(d.style.format({"STR": "{:.2%}", "DOC Index": "{:.1f}"}),
                     hide_index=True, use_container_width=True)


# ──────────────────────────────────────────────────────────────
# BIGBASKET SECTION
# ──────────────────────────────────────────────────────────────

def _render_bigbasket(history_df: pd.DataFrame):
    st.markdown("### 🛒 BigBasket Promo Generator")
    st.caption(
        "Generates a date-range CSV with state columns for BigBasket bulk upload. "
        "Each city is classified into a pricing tier based on sell-through rate and velocity."
    )

    BB_CHANNEL_NAMES = {"BigBasket", "Big Basket", "bigbasket", "big basket"}
    bb_hist = history_df[history_df["channel"].str.strip().isin(BB_CHANNEL_NAMES)].copy()

    if bb_hist.empty:
        st.warning("No BigBasket sales data found. Upload BigBasket data via Smart Upload first.")
        return

    if "city" not in bb_hist.columns or bb_hist["city"].isna().all():
        st.warning("No city-level data for BigBasket. Re-upload with the City/DC column mapped.")
        return

    bb_hist = bb_hist.dropna(subset=["city"])

    # ── Signal / lookback controls ────────────────────────────────────────────
    c1, c2 = st.columns(2)
    with c1:
        lookback = st.slider("Lookback Window (days)", 7, 90, 30, 7, key="bb_lookback")
    with c2:
        doc_thresh = st.slider("DOC Threshold", 30, 90, 70, 5, key="bb_doc_thresh")

    # ── Channel Performance panel ─────────────────────────────────────────────
    _render_channel_performance(bb_hist, "BigBasket", lookback)
    st.divider()

    sig_df = _compute_signals_with_city(bb_hist, lookback)
    if sig_df.empty:
        st.error("Could not compute signals — check that BigBasket data has qty_sold and city.")
        return

    _signals_expander(sig_df, doc_thresh, has_city=True)
    st.divider()

    # ── Step 1: City → BB State Mapping ──────────────────────────────────────
    st.markdown("#### Step 1 — Map Cities to BigBasket State Columns")
    unique_cities = sorted(sig_df["location"].dropna().unique())
    map_data = [{"City (from sales data)": c,
                 "BB State Column": DEFAULT_CITY_MAP.get(c, BB_STATE_COLUMNS[0])}
                for c in unique_cities]
    edited_map_df = st.data_editor(
        pd.DataFrame(map_data),
        column_config={"BB State Column": st.column_config.SelectboxColumn(
            "BB State Column", options=BB_STATE_COLUMNS, required=True)},
        num_rows="fixed", hide_index=True, key="bb_city_map",
    )
    city_to_state = dict(zip(edited_map_df["City (from sales data)"],
                             edited_map_df["BB State Column"]))
    st.divider()

    # ── Step 2: Promo month ───────────────────────────────────────────────────
    st.markdown("#### Step 2 — Select Promo Month")
    cy1, cy2 = st.columns(2)
    with cy1:
        promo_year = st.selectbox("Year", [2025, 2026, 2027], index=1, key="bb_year")
    with cy2:
        promo_month = st.selectbox("Month", list(range(1, 13)),
                                   format_func=lambda x: calendar.month_name[x], key="bb_month")
    st.divider()

    # ── Step 3: Prices ────────────────────────────────────────────────────────
    st.markdown("#### Step 3 — Enter Target Prices per SKU × Tier")
    st.caption("Enter the fixed price (₹) per tier. Leave 0 to skip that tier.")

    unique_skus = sorted(sig_df["channel_sku"].unique())
    h_cols = st.columns([2, 1, 1, 1, 1])
    h_cols[0].markdown("**SKU**")
    for i, tier in enumerate(TIER_LABELS):
        h_cols[i + 1].markdown(f"**{tier}**", help=TIER_HELP[tier])

    sku_prices: dict = {}
    for sku in unique_skus:
        rc = st.columns([2, 1, 1, 1, 1])
        rc[0].write(sku)
        sku_prices[sku] = {
            tier: rc[i + 1].number_input(
                tier, key=f"bb_price_{sku}_{tier}", label_visibility="collapsed",
                min_value=0.0, step=0.5, format="%.2f")
            for i, tier in enumerate(TIER_LABELS)
        }
    st.divider()

    # ── Generate ──────────────────────────────────────────────────────────────
    if st.button("🚀 Generate BigBasket Promo File", type="primary", key="bb_generate"):
        num_days = calendar.monthrange(promo_year, promo_month)[1]
        start_dt = date(promo_year, promo_month, 1)
        end_dt   = date(promo_year, promo_month, num_days)
        sig_lkp  = {(r["channel_sku"], r["location"]): r for _, r in sig_df.iterrows()}
        final_rows = []

        for sku in unique_skus:
            sku_cities = sig_df[sig_df["channel_sku"] == sku]["location"].unique()
            day_by_day, curr = [], start_dt
            while curr <= end_dt:
                is_svd  = curr.day <= 10
                is_wknd = curr.weekday() >= 5
                tier_states: dict = {t: [] for t in TIER_LABELS}
                for city in sku_cities:
                    key = (sku, city)
                    if key not in sig_lkp:
                        continue
                    r = sig_lkp[key]
                    tier  = _classify(r["str"], r["doc"], is_svd, is_wknd, doc_thresh)
                    state = city_to_state.get(city)
                    if state:
                        tier_states[tier].append(state)
                for tier, states in tier_states.items():
                    if states and sku_prices[sku][tier] > 0:
                        day_by_day.append({"Date": curr, "Tier": tier,
                                           "Price": sku_prices[sku][tier],
                                           "States": sorted(set(states))})
                curr += timedelta(days=1)

            if not day_by_day:
                continue

            tier_days: dict = {}
            for entry in day_by_day:
                tier_days.setdefault(entry["Tier"], []).append(entry)

            for tier, entries in tier_days.items():
                entries = sorted(entries, key=lambda x: x["Date"])
                grouped, seg, prev = [], entries[0], entries[0]
                for i in range(1, len(entries)):
                    nxt = entries[i]
                    if (nxt["Price"] != prev["Price"] or nxt["States"] != prev["States"] or
                            nxt["Date"] != prev["Date"] + timedelta(days=1)):
                        grouped.append({"s": seg["Date"], "e": prev["Date"],
                                        "p": prev["Price"], "st": prev["States"]})
                        seg = nxt
                    prev = nxt
                grouped.append({"s": seg["Date"], "e": entries[-1]["Date"],
                                "p": entries[-1]["Price"], "st": entries[-1]["States"]})
                for g in grouped:
                    out = {col: "" for col in BB_HEADERS}
                    out.update({"Code": sku, "Product Description": sku,
                                "Start Date (DD-MM-YYYY)": g["s"].strftime("%d-%m-%Y"),
                                "End Date (DD-MM-YYYY)":   g["e"].strftime("%d-%m-%Y"),
                                "Discount Type": "fixed", "Discount Value": g["p"],
                                "Pan India": "No"})
                    for s in g["st"]:
                        if s in out:
                            out[s] = "Yes"
                    final_rows.append(out)

        if not final_rows:
            st.warning("No promo rows generated. Check that prices are non-zero and signals "
                       "qualify cities into a tier.")
            return

        output_df = pd.DataFrame(final_rows, columns=BB_HEADERS)
        st.success(f"✅ Generated **{len(output_df)}** promo lines for BigBasket!")
        with st.expander("📋 Preview", expanded=True):
            st.dataframe(output_df, hide_index=True, use_container_width=True)
        csv_bytes = output_df.to_csv(index=False).encode("utf-8")
        st.download_button(
            label="⬇️ Download BigBasket Promo CSV",
            data=csv_bytes,
            file_name=f"BB_Promo_{calendar.month_abbr[promo_month]}_{promo_year}.csv",
            mime="text/csv", type="primary",
        )


# ──────────────────────────────────────────────────────────────
# AMAZON SECTION
# ──────────────────────────────────────────────────────────────

def _render_amazon(history_df: pd.DataFrame):
    st.markdown("### 📦 Amazon Deal Sheet Filler")
    st.caption(
        "SVD PRICE and BAU PRICE are filled into the RK WORLD deal sheet template "
        "using performance signals from your sales data. "
        "All original formulas and formatting are preserved."
    )

    AMZN_CHANNEL_NAMES = {"Amazon", "amazon", "Amazon.in", "amazon.in"}
    amzn_hist     = history_df[history_df["channel"].str.strip().isin(AMZN_CHANNEL_NAMES)].copy()
    has_amzn_data = not amzn_hist.empty

    # ── Signal controls ───────────────────────────────────────────────────────
    c1, c2 = st.columns(2)
    with c1:
        lookback = st.slider("Lookback Window (days)", 7, 90, 30, 7, key="amzn_lookback")
    with c2:
        doc_thresh = st.slider("DOC Threshold", 30, 90, 70, 5, key="amzn_doc_thresh")

    # ── Channel Performance panel ─────────────────────────────────────────────
    if has_amzn_data:
        _render_channel_performance(amzn_hist, "Amazon", lookback)
    else:
        st.info("No Amazon sales data found in the database. "
                "Upload Amazon data via Smart Upload to see performance signals. "
                "Prices can still be entered manually below.")
    st.divider()

    # ── Signals ───────────────────────────────────────────────────────────────
    sig_df = pd.DataFrame(columns=["channel_sku", "str", "doc"])
    if has_amzn_data:
        sig_df = _compute_signals_national(amzn_hist, lookback)
        if not sig_df.empty:
            _signals_expander(sig_df, doc_thresh, has_city=False)

    sig_lookup = {r["channel_sku"]: r for _, r in sig_df.iterrows()} if not sig_df.empty else {}
    st.divider()

    # ── Load embedded template ────────────────────────────────────────────────
    try:
        wb_check = load_workbook(io.BytesIO(AMZN_TEMPLATE_BYTES))
        ws_check = wb_check.active
    except Exception as e:
        st.error(f"Could not load embedded template: {e}")
        return

    header_row = [ws_check.cell(1, c).value for c in range(1, ws_check.max_column + 1)]
    data_rows  = []
    for row in ws_check.iter_rows(min_row=AMZN_DATA_START_ROW, values_only=True):
        if any(v is not None for v in row):
            data_rows.append(dict(zip(header_row, row)))

    template_df = pd.DataFrame(data_rows).dropna(subset=["ASIN"])

    with st.expander("📄 Deal Sheet Template (8 ASINs pre-loaded)", expanded=False):
        cols_to_show = [c for c in ["ASIN", "Title", "MRP", "ASP", "Margin",
                                    "SVD PRICE", "BAU PRICE"] if c in template_df.columns]
        st.dataframe(template_df[cols_to_show], hide_index=True, use_container_width=True)

    st.divider()

    # ── Map ASINs → Master SKUs ───────────────────────────────────────────────
    st.markdown("#### Step 1 — Map ASINs to Master SKUs")
    st.caption(
        "Match each ASIN to the master SKU from your sales database. "
        "This links performance signals to the right row."
    )

    master_skus_in_history = (
        sorted(history_df["item_name"].dropna().unique()) if not history_df.empty else []
    )
    sku_options = ["— skip this row —"] + master_skus_in_history

    asin_to_master: dict = {}
    for _, trow in template_df.iterrows():
        asin  = str(trow.get("ASIN", "")).strip()
        title = str(trow.get("Title", ""))[:80]
        auto_match  = next(
            (s for s in master_skus_in_history if s.lower() in title.lower()), None)
        default_idx = sku_options.index(auto_match) if auto_match else 0
        asin_to_master[asin] = st.selectbox(
            f"`{asin}` — {title[:70]}…",
            sku_options, index=default_idx, key=f"amzn_map_{asin}",
        )

    st.divider()

    # ── Review & confirm prices ───────────────────────────────────────────────
    st.markdown("#### Step 2 — Review & Confirm Prices")
    st.caption(
        "Prices are auto-suggested from performance signals. "
        "🟡 SVD = slow-mover deal (days 1–10) · 🟢 BAU = everyday price. "
        "Override any value before generating."
    )

    TIER_COLOUR  = {"SVD": "🟡", "Liq": "🔴", "BAU": "🟢", "Weekend": "🔵"}
    SVD_DISCOUNT = 0.08

    h = st.columns([1.2, 2.8, 0.8, 0.8, 1, 1])
    for lbl, col in zip(["ASIN", "Title", "ASP", "Tier", "SVD Price ✏️", "BAU Price ✏️"], h):
        col.markdown(f"**{lbl}**")

    final_prices: dict = {}
    for _, trow in template_df.iterrows():
        asin   = str(trow.get("ASIN", "")).strip()
        title  = str(trow.get("Title", ""))[:65]
        asp    = float(trow.get("ASP") or 0)
        master = asin_to_master.get(asin, "— skip this row —")

        tier = "BAU"
        if master != "— skip this row —" and master in sig_lookup:
            sig  = sig_lookup[master]
            tier = _classify(sig["str"], sig["doc"], True, False, doc_thresh)

        svd_suggest = round(asp * (1 - SVD_DISCOUNT)) if tier in ("SVD", "Liq") else asp
        bau_suggest = asp

        rc = st.columns([1.2, 2.8, 0.8, 0.8, 1, 1])
        rc[0].caption(asin)
        rc[1].caption(title + "…")
        rc[2].write(f"₹{asp:.0f}")
        rc[3].write(f"{TIER_COLOUR.get(tier, '')} {tier}")

        svd_price = rc[4].number_input(
            "SVD", key=f"amzn_svd_{asin}", value=float(svd_suggest),
            min_value=0.0, step=1.0, format="%.0f", label_visibility="collapsed",
        )
        bau_price = rc[5].number_input(
            "BAU", key=f"amzn_bau_{asin}", value=float(bau_suggest),
            min_value=0.0, step=1.0, format="%.0f", label_visibility="collapsed",
        )
        final_prices[asin] = {"svd": svd_price, "bau": bau_price, "master": master}

    st.divider()

    # ── Generate ──────────────────────────────────────────────────────────────
    if st.button("🚀 Generate Filled Amazon Deal Sheet", type="primary", key="amzn_generate"):
        wb_out = load_workbook(io.BytesIO(AMZN_TEMPLATE_BYTES))
        ws_out = wb_out.active
        blue_font = Font(color="0000FF")
        filled_count, skipped = 0, []

        for row_idx in range(AMZN_DATA_START_ROW, ws_out.max_row + 1):
            asin_val = str(ws_out.cell(row_idx, 1).value or "").strip()
            if not asin_val or asin_val not in final_prices:
                continue
            prices = final_prices[asin_val]
            if prices["master"] == "— skip this row —":
                skipped.append(asin_val)
                continue

            for col_idx, val in [(AMZN_COL_SVD, prices["svd"]),
                                  (AMZN_COL_BAU, prices["bau"])]:
                cell = ws_out.cell(row_idx, col_idx)
                cell.value         = val
                cell.font          = blue_font
                cell.number_format = "#,##0"

            filled_count += 1

        if filled_count == 0:
            st.warning("No rows filled. Map at least one ASIN to a master SKU and try again.")
            return

        out_buf = io.BytesIO()
        wb_out.save(out_buf)
        out_buf.seek(0)

        st.success(f"✅ Filled prices for **{filled_count}** ASINs.")
        if skipped:
            st.info(f"Skipped {len(skipped)} unmapped rows: {', '.join(skipped)}")

        preview_rows = []
        for _, trow in template_df.iterrows():
            asin = str(trow.get("ASIN", "")).strip()
            if asin in final_prices and final_prices[asin]["master"] != "— skip this row —":
                preview_rows.append({
                    "ASIN":          asin,
                    "Title":         str(trow.get("Title", ""))[:70],
                    "ASP (₹)":       trow.get("ASP"),
                    "Mapped SKU":    final_prices[asin]["master"],
                    "SVD PRICE ✏️": final_prices[asin]["svd"],
                    "BAU PRICE ✏️": final_prices[asin]["bau"],
                })

        with st.expander("📋 Filled Price Preview", expanded=True):
            st.dataframe(pd.DataFrame(preview_rows), hide_index=True, use_container_width=True)

        st.download_button(
            label="⬇️ Download Filled Amazon Deal Sheet (.xlsx)",
            data=out_buf,
            file_name=f"Amazon_DealSheet_Filled_{date.today().strftime('%b%Y')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
        )


# ──────────────────────────────────────────────────────────────
# MAIN ENTRY POINT
# ──────────────────────────────────────────────────────────────

def render_deals_promos_tab(history_df: pd.DataFrame, role: str):
    st.subheader("🏷️ Deals & Promos Generator")
    st.caption(
        "Uses channel performance signals from your historical sales data to "
        "auto-suggest pricing tiers and generate the channel-specific deal file."
    )

    channel = st.selectbox(
        "Select Channel", ["BigBasket", "Amazon"],
        help="BigBasket → date-range promo CSV. Amazon → filled deal sheet (.xlsx).",
        key="deals_channel_select",
    )
    st.divider()

    if channel == "BigBasket":
        _render_bigbasket(history_df)
    elif channel == "Amazon":
        _render_amazon(history_df)
